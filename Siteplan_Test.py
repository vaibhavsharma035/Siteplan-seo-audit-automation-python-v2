import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import random
import threading
import keyboard
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
import urllib3

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==========================================================
# CONFIGURATION
# ==========================================================

INPUT_FILE  = "SiteplanSheetName.xlsx"
SHEET_NAME  = "Ops Center"

REQUIRED_COLUMNS = [
    "42Works - Staging Site Link",
    "New H1",
    "Live Title Tag",
    "Live Meta Description"
]

REQUEST_TIMEOUT = 15
DELAY_BETWEEN_REQUESTS = 1

# ── Retry / cache-buster settings ─────────────────────────
MAX_RETRIES = 3
RETRY_BACKOFF_BASE = 4  # seconds, multiplied by attempt number


# ==========================================================
# COLORS  (only non-PASS items get colored)
# ==========================================================

# Red — hard failures
FILL_FAIL      = PatternFill(fill_type="solid", fgColor="FFC7CE")
FONT_FAIL      = Font(color="9C0006", bold=True)

# Orange — warnings (multiple H1, missing, no expected)
FILL_WARN      = PatternFill(fill_type="solid", fgColor="FFEB9C")
FONT_WARN      = Font(color="9C5700", bold=False)

# Purple — 4xx / 5xx status codes
FILL_HTTP_ERR  = PatternFill(fill_type="solid", fgColor="E2CFFE")
FONT_HTTP_ERR  = Font(color="4B0082", bold=True)

# Grey — request errors (timeout, SSL, connection)
FILL_ERROR     = PatternFill(fill_type="solid", fgColor="D9D9D9")
FONT_ERROR     = Font(color="595959", bold=False)

# Which result values get which style
FAIL_VALUES    = {"FAIL"}
WARN_VALUES    = {"MULTIPLE H1", "H1 MISSING", "TITLE MISSING", "META MISSING",
                  "NO EXPECTED H1", "NO EXPECTED TITLE", "NO EXPECTED META"}
HTTP_ERR_CODES = {"404", "403", "500", "502", "503", "301", "302"}


def apply_cell_style(cell, value):
    """
    Apply fill + font to a single cell based on its value.
    PASS and empty cells are left untouched.
    """
    val = str(value).strip().upper()

    if val in FAIL_VALUES:
        cell.fill = FILL_FAIL
        cell.font = FONT_FAIL

    elif val in WARN_VALUES:
        cell.fill = FILL_WARN
        cell.font = FONT_WARN

    elif val == "ERROR":
        cell.fill = FILL_ERROR
        cell.font = FONT_ERROR


def apply_status_style(cell, value):
    """
    Apply style to the Status_Code cell.
    4xx / 5xx / ERROR get colored. 200 is left alone.
    """
    val = str(value).strip()

    if val == "ERROR":
        cell.fill = FILL_ERROR
        cell.font = FONT_ERROR

    elif val in HTTP_ERR_CODES or (val.isdigit() and int(val) >= 400):
        cell.fill = FILL_HTTP_ERR
        cell.font = FONT_HTTP_ERR


# ==========================================================
# ESC LISTENER
# ==========================================================

stop_execution = False


def esc_listener():

    global stop_execution

    keyboard.wait("esc")

    print("\n" + "=" * 80)
    print("ESC pressed.")
    print("Finishing current URL and stopping...")
    print("=" * 80)

    stop_execution = True


threading.Thread(
    target=esc_listener,
    daemon=True
).start()


# ==========================================================
# HELPER — clean text
# ==========================================================

def clean_text(text):

    if text is None or pd.isna(text):
        return ""

    return " ".join(
        str(text)
        .replace("\n", " ")
        .replace("\r", " ")
        .split()
    ).strip()


# ==========================================================
# HELPER — cache-busting + retry logic
# ==========================================================

def add_cache_buster(url):
    """
    Append a unique query param so CDNs/proxies/browser-level caches
    don't serve a stale copy of the page. Preserves any existing
    query string on the URL.
    """
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query))
    query["_cb"] = str(int(time.time() * 1000)) + str(random.randint(1000, 9999))
    new_query = urlencode(query)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))


def fetch_with_retry(session, url, timeout, max_retries=MAX_RETRIES):
    """
    Fetches a URL with cache-busting and retry logic.

    - Adds a unique query string on every attempt to bypass CDN /
      reverse-proxy caching (fixes stale-cache FAILs).
    - Retries on 429 (rate limit), respecting Retry-After if present.
    - Retries on 404, since some pages briefly 404 during redirects
      or WAF challenges even though the page is actually live.
    - Retries on timeout / SSL / connection errors with backoff.

    Returns (response, None) on success, or (None, exception) on
    exhausted retries.
    """
    last_exception = None

    for attempt in range(1, max_retries + 1):

        busted_url = add_cache_buster(url)

        try:
            response = session.get(
                busted_url,
                timeout=timeout,
                verify=False,
                allow_redirects=True
            )

            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                if retry_after and retry_after.isdigit():
                    wait_time = int(retry_after)
                else:
                    wait_time = RETRY_BACKOFF_BASE * attempt
                print(f"  429 received. Waiting {wait_time}s before retry {attempt}/{max_retries}...")
                time.sleep(wait_time)
                continue

            if response.status_code == 404 and attempt < max_retries:
                wait_time = RETRY_BACKOFF_BASE * attempt
                print(f"  404 received. Retrying in {wait_time}s ({attempt}/{max_retries})...")
                time.sleep(wait_time)
                continue

            return response, None

        except (requests.exceptions.Timeout,
                requests.exceptions.SSLError,
                requests.exceptions.ConnectionError) as e:

            last_exception = e
            wait_time = RETRY_BACKOFF_BASE * attempt
            print(f"  Request error ({type(e).__name__}). Retrying in {wait_time}s ({attempt}/{max_retries})...")
            time.sleep(wait_time)

    return None, last_exception


# ==========================================================
# LOAD DATA
# ==========================================================

print("\n" + "=" * 80)
print("SEO Checker")
print("=" * 80)
print(f"Input File  : {INPUT_FILE}")
print(f"Sheet       : {SHEET_NAME}")
print("=" * 80)


def find_header_row(file_path, sheet_name, required_columns, max_scan_rows=10):
    """
    Scans the first `max_scan_rows` rows of the sheet (with no header
    assumed) to find which row actually contains all of REQUIRED_COLUMNS.
    Returns the 0-indexed row number to pass as header= to read_excel,
    or None if no matching row was found.
    """
    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_scan_rows)

    for row_idx in range(len(raw)):
        row_values = raw.iloc[row_idx].astype(str).str.strip().tolist()
        if all(col in row_values for col in required_columns):
            return row_idx

    return None


header_row = find_header_row(INPUT_FILE, SHEET_NAME, REQUIRED_COLUMNS)

if header_row is None:
    # Couldn't find it automatically — show the first several rows so the
    # user can see the exact header text and fix REQUIRED_COLUMNS.
    print("\nCould NOT auto-detect the header row. Here are the first 10 rows of the sheet:")
    preview = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=None, nrows=10)
    for i, r in preview.iterrows():
        print(f"Row {i}: {r.tolist()}")
    raise SystemExit(
        "\nUpdate REQUIRED_COLUMNS to match the exact header text shown above, "
        "or set the header row manually."
    )

print(f"\nDetected header row at Excel row index: {header_row}")

df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=header_row)

df.columns = df.columns.astype(str).str.strip()

print("\n=== FOUND COLUMNS ===")
print(df.columns.tolist())

df = df[REQUIRED_COLUMNS]

df.columns = ["URL", "Expected_H1", "Expected_Title", "Expected_Meta"]

df = df[df["URL"].notna()].reset_index(drop=True)

total_urls = len(df)

print(f"\nTotal URLs to check : {total_urls}")
print("Press ESC anytime to stop safely.")
print("=" * 80)


# ==========================================================
# RESULT COLUMNS
# ==========================================================

df["Status_Code"]   = ""
df["Actual_H1"]     = ""
df["H1_Count"]      = 0
df["H1_Result"]     = ""
df["Actual_Title"]  = ""
df["Title_Result"]  = ""
df["Actual_Meta"]   = ""
df["Meta_Result"]   = ""
df["Checked_At"]    = ""


# ==========================================================
# REQUEST SESSION
# ==========================================================

session = requests.Session()

session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Cache-Control": "no-cache, no-store, must-revalidate",
    "Pragma": "no-cache"
})


# ==========================================================
# COUNTERS
# ==========================================================

pass_count  = 0
fail_count  = 0
warn_count  = 0
error_count = 0


# ==========================================================
# MAIN LOOP
# ==========================================================

for index, row in df.iterrows():

    if stop_execution:
        print("\nStopped by user. Saving results...")
        break

    url            = row["URL"]
    expected_h1    = clean_text(row["Expected_H1"])
    expected_title = clean_text(row["Expected_Title"])
    expected_meta  = clean_text(row["Expected_Meta"])

    print("\n" + "-" * 80)
    print(f"Processing : {index + 1} / {total_urls}")
    print(f"URL        : {url}")

    try:

        response, fetch_error = fetch_with_retry(session, url, REQUEST_TIMEOUT)

        if fetch_error is not None:
            raise fetch_error

        if response is None:
            raise Exception("Max retries exceeded")

        status_code = response.status_code

        soup = BeautifulSoup(response.text, "html.parser")

        # ── H1 ──────────────────────────────────────────
        h1_tags = soup.find_all("h1")
        h1_count = len(h1_tags)
        actual_h1 = (
            clean_text(h1_tags[0].get_text())
            if h1_count > 0 else ""
        )

        # ── Title ────────────────────────────────────────
        actual_title = (
            clean_text(soup.title.get_text())
            if soup.title else ""
        )

        # ── Meta Description ─────────────────────────────
        meta_tag = soup.find("meta", attrs={"name": "description"})

        if not meta_tag:
            meta_tag = soup.find("meta", attrs={"property": "og:description"})

        actual_meta = (
            clean_text(meta_tag.get("content"))
            if meta_tag and meta_tag.get("content")
            else ""
        )

    except requests.exceptions.Timeout:

        print(f"TIMEOUT : {url}")

        status_code  = "ERROR"
        actual_h1    = ""
        actual_title = ""
        actual_meta  = ""
        h1_count     = 0

    except requests.exceptions.SSLError:

        print(f"SSL ERROR : {url}")

        status_code  = "ERROR"
        actual_h1    = ""
        actual_title = ""
        actual_meta  = ""
        h1_count     = 0

    except Exception as e:

        print(f"ERROR : {url} | {e}")

        status_code  = "ERROR"
        actual_h1    = ""
        actual_title = ""
        actual_meta  = ""
        h1_count     = 0

    # ── Save raw values ──────────────────────────────────
    df.at[index, "Status_Code"]  = str(status_code)
    df.at[index, "Actual_H1"]    = actual_h1
    df.at[index, "H1_Count"]     = h1_count
    df.at[index, "Actual_Title"] = actual_title
    df.at[index, "Actual_Meta"]  = actual_meta
    df.at[index, "Checked_At"]   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── H1 Result ────────────────────────────────────────
    if status_code == "ERROR":
        h1_result = "ERROR"

    elif expected_h1 == "":
        h1_result = "NO EXPECTED H1"

    elif actual_h1 == "":
        h1_result = "H1 MISSING"

    elif h1_count > 1:
        h1_result = "MULTIPLE H1"

    elif expected_h1.lower() == actual_h1.lower():
        h1_result = "PASS"

    else:
        h1_result = "FAIL"

    # ── Title Result ─────────────────────────────────────
    if status_code == "ERROR":
        title_result = "ERROR"

    elif expected_title == "":
        title_result = "NO EXPECTED TITLE"

    elif actual_title == "":
        title_result = "TITLE MISSING"

    elif expected_title.lower() == actual_title.lower():
        title_result = "PASS"

    else:
        title_result = "FAIL"

    # ── Meta Result ──────────────────────────────────────
    if status_code == "ERROR":
        meta_result = "META ERROR"

    elif expected_meta == "":
        meta_result = "NO EXPECTED META"

    elif actual_meta == "":
        meta_result = "META MISSING"

    elif expected_meta.lower() == actual_meta.lower():
        meta_result = "PASS"

    else:
        meta_result = "FAIL"

    # ── Save results ─────────────────────────────────────
    df.at[index, "H1_Result"]    = h1_result
    df.at[index, "Title_Result"] = title_result
    df.at[index, "Meta_Result"]  = meta_result

    # ── Count ────────────────────────────────────────────
    row_results = [h1_result, title_result, meta_result]

    if "FAIL" in row_results:
        fail_count += 1
    elif "ERROR" in " ".join(row_results) or status_code == "ERROR":
        error_count += 1
    elif any(r in WARN_VALUES for r in row_results):
        warn_count += 1
    else:
        pass_count += 1

    print(f"Status     : {status_code}")
    print(f"H1 Result  : {h1_result}  (count: {h1_count})")
    print(f"Title      : {title_result}")
    print(f"Meta       : {meta_result}")

    time.sleep(DELAY_BETWEEN_REQUESTS)


# ==========================================================
# SAVE OUTPUT FILE
# ==========================================================

source_name = Path(INPUT_FILE).stem

number = 1

while True:

    output_file = f"{source_name}_SEO_Result_{number:03d}.xlsx"

    if not Path(output_file).exists():
        break

    number += 1

df.to_excel(output_file, index=False)

print(f"\nBase file saved : {output_file}")


# ==========================================================
# APPLY FORMATTING
# ==========================================================

wb = load_workbook(output_file)
ws = wb.active

# ── Build column index map from header row ───────────────
col_map = {}

for cell in ws[1]:
    if cell.value:
        col_map[str(cell.value).strip()] = cell.column

result_columns = ["H1_Result", "Title_Result", "Meta_Result"]
status_column  = "Status_Code"

# ── Header row formatting ────────────────────────────────
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")  # dark navy blue
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)           # white bold text
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="medium", color="FFFFFF")
)

for cell in ws[1]:

    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = HEADER_BORDER

# ── Freeze the header row (sticky) ──────────────────────
ws.freeze_panes = "A2"

# ── Set header row height ────────────────────────────────
ws.row_dimensions[1].height = 35

# ── Auto column width ────────────────────────────────────
for col in ws.columns:

    max_length = 0

    col_letter = col[0].column_letter

    for cell in col:

        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except Exception:
            pass

    # Cap width at 60, minimum 12
    adjusted_width = min(max(max_length + 2, 3), 16)

    ws.column_dimensions[col_letter].width = adjusted_width

# ── Color result and status cells in data rows ───────────
for row in range(2, ws.max_row + 1):

    # Color result columns
    for col_name in result_columns:

        if col_name in col_map:

            cell = ws.cell(
                row=row,
                column=col_map[col_name]
            )

            apply_cell_style(cell, cell.value)

    # Color status code column
    if status_column in col_map:

        cell = ws.cell(
            row=row,
            column=col_map[status_column]
        )

        apply_status_style(cell, cell.value)

wb.save(output_file)

print(f"Formatting applied.")


# ==========================================================
# FINAL SUMMARY
# ==========================================================

processed = index + 1 if not stop_execution else index
remaining = total_urls - processed

if remaining < 0:
    remaining = 0

print("\n")
print("=" * 80)
print("SEO CHECK COMPLETED")
print("=" * 80)
print(f"Input File        : {INPUT_FILE}")
print(f"Sheet             : {SHEET_NAME}")
print("-" * 80)
print(f"Total URLs        : {total_urls}")
print(f"Processed         : {processed}")
print(f"Passed            : {pass_count}")
print(f"Failed            : {fail_count}")
print(f"Warnings          : {warn_count}")
print(f"Errors            : {error_count}")
print(f"Remaining         : {remaining}")
print("-" * 80)

if stop_execution:
    print("Status            : STOPPED BY USER (ESC)")
else:
    print("Status            : COMPLETED")

print("=" * 80)
print(f"Output File       : {output_file}")
print("=" * 80)
