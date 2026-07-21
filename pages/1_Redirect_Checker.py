import io
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import urllib3

# pages/ is a sibling of the repo root where shared_utils.py lives —
# Streamlit runs each page with the repo root already on sys.path, so
# this import works the same as it does in streamlit_app.py.
from shared_utils import (
    fetch_with_retry,
    guess_header_row,
    best_match_index,
    new_requests_session,
    build_data_source_ui,
)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================================================
# PAGE CONFIG
# ==========================================================
st.set_page_config(page_title="Redirect Checker", layout="wide")

# ==========================================================
# DEFAULTS
# ==========================================================
DEFAULT_TIMEOUT = 15
DEFAULT_DELAY = 1
DEFAULT_MAX_RETRIES = 3

SKIP_EXTENSIONS = (
    ".css", ".js", ".png", ".jpg", ".jpeg", ".gif", ".svg", ".ico",
    ".woff", ".woff2", ".ttf", ".eot", ".pdf", ".xml", ".zip", ".rar",
    ".7z", ".txt",
)

# ==========================================================
# ROW COLORS
# ==========================================================
FILL_PASS = PatternFill(fill_type="solid", fgColor="C6EFCE")
FONT_PASS = Font(color="276221", bold=False)

FILL_FAIL = PatternFill(fill_type="solid", fgColor="FFC7CE")
FONT_FAIL = Font(color="9C0006", bold=True)

FILL_ERROR = PatternFill(fill_type="solid", fgColor="FFEB9C")
FONT_ERROR = Font(color="9C5700", bold=False)

FILL_SKIPPED = PatternFill(fill_type="solid", fgColor="D9D9D9")
FONT_SKIPPED = Font(color="595959", bold=False)

ROW_STYLE = {
    "PASS": (FILL_PASS, FONT_PASS),
    "FAIL": (FILL_FAIL, FONT_FAIL),
    "ERROR": (FILL_ERROR, FONT_ERROR),
    "SKIPPED": (FILL_SKIPPED, FONT_SKIPPED),
}


def color_row(ws, row_num, result, result_col_index):
    fill, font = ROW_STYLE.get(result, (None, None))
    if not fill:
        return
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_num, column=col).fill = fill
    ws.cell(row=row_num, column=result_col_index).font = font


def normalize(url):
    if not url:
        return ""
    parsed = urlparse(str(url).strip())
    scheme = parsed.scheme.lower()
    netloc = parsed.netloc.lower()
    path = parsed.path.rstrip("/")
    return f"{scheme}://{netloc}{path}"


def is_asset_url(url):
    return str(url).strip().lower().endswith(SKIP_EXTENSIONS)


# ==========================================================
# UI — HEADER
# ==========================================================
st.title("🔀 Redirect Checker")
st.caption("Load a sheet, map your Live/Staging URL and expected redirect-to URL columns, and download a color-coded results workbook.")

with st.sidebar:
    st.header("Run settings")
    request_timeout = st.number_input("Request timeout (seconds)", min_value=5, max_value=60, value=DEFAULT_TIMEOUT)
    delay_between_requests = st.number_input("Delay between requests (seconds)", min_value=0.0, max_value=10.0, value=float(DEFAULT_DELAY), step=0.5)
    max_retries = st.number_input("Max retries per URL", min_value=1, max_value=5, value=DEFAULT_MAX_RETRIES)
    st.markdown("---")
    st.caption(
        "A run can't be safely paused and resumed — closing or refreshing "
        "this tab stops it, and nothing processed so far will be saved."
    )
    st.markdown("---")
    with st.expander("Skipped file types (static assets)"):
        st.caption(", ".join(SKIP_EXTENSIONS))

# ==========================================================
# DATA SOURCE — file upload OR Google Sheets link (shared_utils)
# ==========================================================
file_bytes, source_name = build_data_source_ui(st, key_prefix="redirect")

# ==========================================================
# SHEET / HEADER / COLUMN MAPPING
# ==========================================================
if file_bytes is not None:

    try:
        workbook_preview = pd.ExcelFile(io.BytesIO(file_bytes))
        sheet_names = workbook_preview.sheet_names
    except Exception as e:
        st.error(f"Could not read this as an Excel workbook: {e}")
        st.stop()

    default_sheet_index = sheet_names.index("redirect file") if "redirect file" in sheet_names else 0
    sheet_name = st.selectbox("Sheet", options=sheet_names, index=default_sheet_index)

    # No fixed column-name guesses for this tool (unlike the SEO checker),
    # so header-row detection falls back to row 0 — adjust manually if needed.
    auto_row = guess_header_row(file_bytes, sheet_name, [])
    header_row = st.number_input(
        "Header row (0 = first row of the sheet)",
        min_value=0, max_value=20, value=auto_row,
        help="The row number where your column titles actually appear."
    )

    try:
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row)
        df_raw.columns = df_raw.columns.astype(str).str.strip()
    except Exception as e:
        st.error(f"Could not read the sheet with that header row: {e}")
        st.stop()

    available_columns = df_raw.columns.tolist()

    with st.expander("Preview of detected columns and first rows", expanded=False):
        st.write("Columns found:", available_columns)
        st.dataframe(df_raw.head(5))

    st.subheader("Map your columns")
    st.caption("Live/Staging link is the URL to open. Redirect-to link is where it's expected to land.")

    map_col1, map_col2 = st.columns(2)

    with map_col1:
        live_column = st.selectbox(
            "Live / Staging URL column",
            options=available_columns,
            index=best_match_index(available_columns, "Live URL")
        )

    with map_col2:
        expected_column = st.selectbox(
            "Expected redirect-to URL column",
            options=available_columns,
            index=best_match_index(available_columns, "Redirect To")
        )

    mapping_has_duplicates = live_column == expected_column
    if mapping_has_duplicates:
        st.warning("Both fields point to the same column — double check your selections.")

    run_clicked = st.button("▶ Run Redirect Check", type="primary", disabled=mapping_has_duplicates)

else:
    run_clicked = False
    st.info("Upload an Excel file or paste a Google Sheets link to get started.")

# ==========================================================
# MAIN EXECUTION
# ==========================================================
if file_bytes is not None and run_clicked:

    df = df_raw[[live_column, expected_column]].copy()
    df.columns = ["Live_URL", "Expected_URL"]
    df = df[df["Expected_URL"].notna()].reset_index(drop=True)

    total_urls = len(df)

    st.code(
        f"Input      : {source_name}\n"
        f"Sheet      : {sheet_name}\n"
        f"Total URLs : {total_urls}"
    )

    # This box only ever shows retry activity for the URL currently being
    # processed — reset each iteration and cleared once that row's
    # outcome is known, so old retry messages don't pile up on screen.
    log_box = st.empty()

    df["Actual_URL"] = ""
    df["Status_Code"] = ""
    df["Result"] = ""
    df["Remarks"] = ""
    df["Checked_At"] = ""

    session = new_requests_session()

    pass_count = fail_count = error_count = skipped_count = 0

    progress_bar = st.progress(0.0)
    detail_placeholder = st.empty()

    for index, row in df.iterrows():

        live_url = "" if pd.isna(row["Live_URL"]) else str(row["Live_URL"]).strip()
        expected_url = str(row["Expected_URL"]).strip()

        # Fresh retry log for this URL only.
        log_lines = []

        def log(msg):
            log_lines.append(msg)
            log_box.code("\n".join(log_lines[-40:]))

        detail_placeholder.code(f"Processing : {index + 1} / {total_urls}\nLive URL   : {live_url}")

        if live_url == "":
            result, remarks = "SKIPPED", "Live URL Empty"
            actual_url = ""
            status_code = ""
            skipped_count += 1

        elif is_asset_url(live_url):
            result, remarks = "SKIPPED", "Static Asset"
            actual_url = ""
            status_code = ""
            skipped_count += 1

        else:
            try:
                response, fetch_error = fetch_with_retry(session, live_url, request_timeout, max_retries, log)

                if fetch_error is not None:
                    raise fetch_error
                if response is None:
                    raise Exception("Max retries exceeded")

                status_code = response.status_code
                actual_url = response.url

                if status_code >= 400:
                    # Retries were exhausted and the page still isn't
                    # loading successfully — this is a failure regardless
                    # of what the final URL string happens to be. Without
                    # this check, a 404 page that doesn't redirect
                    # anywhere could be marked PASS just because its URL
                    # matches the expected one.
                    result, remarks = "FAIL", f"Live URL returned HTTP {status_code}"
                    fail_count += 1
                elif normalize(actual_url) == normalize(expected_url):
                    result, remarks = "PASS", "URLs Match"
                    pass_count += 1
                else:
                    result, remarks = "FAIL", "URL Mismatch"
                    fail_count += 1

            except requests.exceptions.Timeout:
                log(f"TIMEOUT : {live_url}")
                result, remarks, actual_url, status_code = "ERROR", "Timeout", "", "ERROR"
                error_count += 1

            except requests.exceptions.SSLError:
                log(f"SSL ERROR : {live_url}")
                result, remarks, actual_url, status_code = "ERROR", "SSL Error", "", "ERROR"
                error_count += 1

            except Exception as e:
                log(f"ERROR : {live_url} | {e}")
                result, remarks, actual_url, status_code = "ERROR", str(e), "", "ERROR"
                error_count += 1

        # This URL is done — clear the transient retry log now that its
        # outcome is captured below and in the final report.
        if log_lines:
            log_box.empty()

        df.at[index, "Actual_URL"] = actual_url
        df.at[index, "Status_Code"] = str(status_code)
        df.at[index, "Result"] = result
        df.at[index, "Remarks"] = remarks
        df.at[index, "Checked_At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        detail_placeholder.code(
            f"Processing : {index + 1} / {total_urls}\n"
            f"Live URL   : {live_url}\n"
            f"Expected   : {expected_url}\n"
            f"Actual     : {actual_url}\n"
            f"Status     : {status_code}\n"
            f"Result     : {result}  ({remarks})"
        )

        progress_bar.progress((index + 1) / total_urls)
        time.sleep(delay_between_requests)

    detail_placeholder.code("Done processing all URLs.")

    # ==========================================================
    # BUILD FORMATTED OUTPUT WORKBOOK (in memory, row-level color)
    # ==========================================================
    excel_buffer = io.BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    wb = load_workbook(excel_buffer)
    ws = wb.active

    col_map = {}
    for cell in ws[1]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    result_col_index = col_map.get("Result")

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HEADER_BORDER = Border(bottom=Side(style="medium", color="FFFFFF"))

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 3), 16)

    if result_col_index:
        for row_num in range(2, ws.max_row + 1):
            result_value = ws.cell(row=row_num, column=result_col_index).value
            color_row(ws, row_num, result_value, result_col_index)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    # ==========================================================
    # SUMMARY + DOWNLOAD
    # ==========================================================
    st.success("Redirect check completed.")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Passed", pass_count)
    c2.metric("Failed", fail_count)
    c3.metric("Errors", error_count)
    c4.metric("Skipped", skipped_count)

    output_filename = f"{source_name}_Redirect_Result_001.xlsx"

    st.download_button(
        label="⬇ Download Results (.xlsx)",
        data=output_buffer,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with st.expander("View results table"):
        st.dataframe(df)
