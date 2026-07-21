import io
import time
from datetime import datetime

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import urllib3

from shared_utils import (
    fetch_with_retry,
    guess_header_row,
    best_match_index,
    new_requests_session,
    build_data_source_ui,
)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================================================
# PAGE CONFIG
# ==========================================================
st.set_page_config(page_title="Siteplan SEO Checker", layout="wide")

# ==========================================================
# DEFAULTS
# ==========================================================
# Used only as a best-guess when auto-selecting dropdown defaults below —
# not a hard requirement anymore, since columns are now chosen in the UI.
DEFAULT_COLUMN_GUESSES = {
    "url": "42Works - Staging Site Link",
    "h1": "New H1",
    "title": "Live Title Tag",
    "meta": "Live Meta Description",
}

DEFAULT_TIMEOUT = 15
DEFAULT_DELAY = 1
DEFAULT_MAX_RETRIES = 3

# ==========================================================
# COLORS
# ==========================================================
FILL_FAIL = PatternFill(fill_type="solid", fgColor="FFC7CE")
FONT_FAIL = Font(color="9C0006", bold=True)

FILL_WARN = PatternFill(fill_type="solid", fgColor="FFEB9C")
FONT_WARN = Font(color="9C5700", bold=False)

FILL_HTTP_ERR = PatternFill(fill_type="solid", fgColor="E2CFFE")
FONT_HTTP_ERR = Font(color="4B0082", bold=True)

FILL_ERROR = PatternFill(fill_type="solid", fgColor="D9D9D9")
FONT_ERROR = Font(color="595959", bold=False)

FAIL_VALUES = {"FAIL"}
WARN_VALUES = {"MULTIPLE H1", "H1 MISSING", "TITLE MISSING", "META MISSING",
               "NO EXPECTED H1", "NO EXPECTED TITLE", "NO EXPECTED META"}
HTTP_ERR_CODES = {"404", "403", "500", "502", "503", "301", "302"}
ERROR_VALUES = {"ERROR"}


def apply_cell_style(cell, value):
    val = str(value).strip().upper()
    if val in FAIL_VALUES:
        cell.fill = FILL_FAIL
        cell.font = FONT_FAIL
    elif val in WARN_VALUES:
        cell.fill = FILL_WARN
        cell.font = FONT_WARN
    elif val in ERROR_VALUES:
        cell.fill = FILL_ERROR
        cell.font = FONT_ERROR


def apply_status_style(cell, value):
    val = str(value).strip()
    if val == "ERROR":
        cell.fill = FILL_ERROR
        cell.font = FONT_ERROR
    elif val in HTTP_ERR_CODES or (val.isdigit() and int(val) >= 400):
        cell.fill = FILL_HTTP_ERR
        cell.font = FONT_HTTP_ERR


def clean_text(text):
    if text is None or pd.isna(text):
        return ""
    return " ".join(str(text).replace("\n", " ").replace("\r", " ").split()).strip()


# ==========================================================
# UI — HEADER
# ==========================================================
st.title("🔍 Siteplan SEO Checker")
st.caption("Load a sheet, choose which fields to check, map your columns, and download a color-coded results workbook.")

with st.sidebar:
    st.header("Run settings")
    request_timeout = st.number_input("Request timeout (seconds)", min_value=5, max_value=60, value=DEFAULT_TIMEOUT)
    delay_between_requests = st.number_input("Delay between requests (seconds)", min_value=0.0, max_value=10.0, value=float(DEFAULT_DELAY), step=0.5)
    max_retries = st.number_input("Max retries per URL", min_value=1, max_value=5, value=DEFAULT_MAX_RETRIES)
    st.markdown("---")
    st.caption(
        "A run can't be safely paused and resumed — closing or refreshing "
        "this tab stops it, and nothing processed so far will be saved. "
        "For very large sheets, consider checking them in smaller batches."
    )

# ==========================================================
# DATA SOURCE — file upload OR Google Sheets link (shared_utils)
# ==========================================================
file_bytes, source_name = build_data_source_ui(st, key_prefix="seo")

# ==========================================================
# SHEET / HEADER / COLUMN MAPPING
# ==========================================================
if file_bytes is not None:

    # ── Sheet selection ───────────────────────────────────
    try:
        workbook_preview = pd.ExcelFile(io.BytesIO(file_bytes))
        sheet_names = workbook_preview.sheet_names
    except Exception as e:
        st.error(f"Could not read this as an Excel workbook: {e}")
        st.stop()

    default_sheet_index = sheet_names.index("Ops Center") if "Ops Center" in sheet_names else 0
    sheet_name = st.selectbox("Sheet", options=sheet_names, index=default_sheet_index)

    # ── Header row selection ──────────────────────────────
    auto_row = guess_header_row(file_bytes, sheet_name, list(DEFAULT_COLUMN_GUESSES.values()))
    header_row = st.number_input(
        "Header row (0 = first row of the sheet)",
        min_value=0, max_value=20, value=auto_row,
        help="The row number where your column titles actually appear. Auto-detected as a starting guess — adjust if it looks wrong."
    )

    try:
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row)
        df_raw.columns = df_raw.columns.astype(str).str.strip()
    except Exception as e:
        st.error(f"Could not read the sheet with that header row: {e}")
        st.stop()

    available_columns = df_raw.columns.tolist()

    with st.expander("Preview of detected columns and first rows", expanded=False):
        st.write("Columns found:", available_columns)
        st.dataframe(df_raw.head(5))

    # ── Field selection — choose which fields to check ────
    st.subheader("Which fields do you want to check?")
    field_col1, field_col2, field_col3 = st.columns(3)
    with field_col1:
        check_h1 = st.checkbox("H1", value=True)
    with field_col2:
        check_title = st.checkbox("Title", value=True)
    with field_col3:
        check_meta = st.checkbox("Meta Description", value=True)

    no_fields_selected = not (check_h1 or check_title or check_meta)
    if no_fields_selected:
        st.warning("Select at least one field to check.")

    # ── Column mapping — only shows dropdowns for selected fields ──
    st.subheader("Map your columns")
    st.caption("Tell the checker which column in your sheet holds each piece of data.")

    map_col1, map_col2 = st.columns(2)

    with map_col1:
        url_column = st.selectbox(
            "Staging URL column",
            options=available_columns,
            index=best_match_index(available_columns, DEFAULT_COLUMN_GUESSES["url"])
        )
        h1_column = None
        if check_h1:
            h1_column = st.selectbox(
                "Expected H1 column",
                options=available_columns,
                index=best_match_index(available_columns, DEFAULT_COLUMN_GUESSES["h1"])
            )

    with map_col2:
        title_column = None
        if check_title:
            title_column = st.selectbox(
                "Expected Title column",
                options=available_columns,
                index=best_match_index(available_columns, DEFAULT_COLUMN_GUESSES["title"])
            )
        meta_column = None
        if check_meta:
            meta_column = st.selectbox(
                "Expected Meta Description column",
                options=available_columns,
                index=best_match_index(available_columns, DEFAULT_COLUMN_GUESSES["meta"])
            )

    selected_columns = [c for c in [url_column, h1_column, title_column, meta_column] if c is not None]
    mapping_has_duplicates = len(set(selected_columns)) != len(selected_columns)

    if mapping_has_duplicates:
        st.warning("You've mapped the same column to more than one field — double check your selections.")

    run_disabled = mapping_has_duplicates or no_fields_selected
    run_clicked = st.button("▶ Run SEO Check", type="primary", disabled=run_disabled)

else:
    run_clicked = False
    st.info("Upload an Excel file or paste a Google Sheets link to get started.")

# ==========================================================
# MAIN EXECUTION
# ==========================================================
if file_bytes is not None and run_clicked:

    rename_map = {url_column: "URL"}
    if check_h1:
        rename_map[h1_column] = "Expected_H1"
    if check_title:
        rename_map[title_column] = "Expected_Title"
    if check_meta:
        rename_map[meta_column] = "Expected_Meta"

    df = df_raw[list(rename_map.keys())].copy()
    df.columns = [rename_map[c] for c in df.columns]
    df = df[df["URL"].notna()].reset_index(drop=True)

    total_urls = len(df)

    fields_being_checked = ", ".join(
        f for f, enabled in [("H1", check_h1), ("Title", check_title), ("Meta", check_meta)] if enabled
    )
    st.code(
        f"Input       : {source_name}\n"
        f"Sheet       : {sheet_name}\n"
        f"Fields      : {fields_being_checked}\n"
        f"Total URLs  : {total_urls}"
    )

    # This box only ever shows retry activity for the URL currently being
    # processed — it's reset at the start of each row and cleared once
    # that row's outcome is known, so old retry messages don't pile up
    # on screen (the outcome itself is already reflected in the detail
    # panel below and in the final report).
    log_box = st.empty()

    df["Status_Code"] = ""
    df["Checked_At"] = ""
    if check_h1:
        df["Actual_H1"] = ""
        df["H1_Count"] = 0
        df["H1_Result"] = ""
    if check_title:
        df["Actual_Title"] = ""
        df["Title_Result"] = ""
    if check_meta:
        df["Actual_Meta"] = ""
        df["Meta_Result"] = ""

    session = new_requests_session()

    pass_count = fail_count = warn_count = error_count = 0

    progress_bar = st.progress(0.0)
    # Live per-URL detail panel — mirrors what you'd see in a local
    # terminal run, showing only the fields actually being checked.
    detail_placeholder = st.empty()

    for index, row in df.iterrows():

        url = row["URL"]
        expected_h1 = clean_text(row["Expected_H1"]) if check_h1 else ""
        expected_title = clean_text(row["Expected_Title"]) if check_title else ""
        expected_meta = clean_text(row["Expected_Meta"]) if check_meta else ""

        # Fresh retry log for this URL only.
        log_lines = []

        def log(msg):
            log_lines.append(msg)
            log_box.code("\n".join(log_lines[-40:]))

        detail_placeholder.code(f"Processing : {index + 1} / {total_urls}\nURL        : {url}")

        try:
            response, fetch_error = fetch_with_retry(session, url, request_timeout, max_retries, log)

            if fetch_error is not None:
                raise fetch_error
            if response is None:
                raise Exception("Max retries exceeded")

            status_code = response.status_code
            soup = BeautifulSoup(response.text, "html.parser")

            actual_h1, h1_count = "", 0
            if check_h1:
                h1_tags = soup.find_all("h1")
                h1_count = len(h1_tags)
                actual_h1 = clean_text(h1_tags[0].get_text()) if h1_count > 0 else ""

            actual_title = ""
            if check_title:
                actual_title = clean_text(soup.title.get_text()) if soup.title else ""

            actual_meta = ""
            if check_meta:
                meta_tag = soup.find("meta", attrs={"name": "description"})
                if not meta_tag:
                    meta_tag = soup.find("meta", attrs={"property": "og:description"})
                actual_meta = clean_text(meta_tag.get("content")) if meta_tag and meta_tag.get("content") else ""

        except requests.exceptions.Timeout:
            log(f"TIMEOUT : {url}")
            status_code, actual_h1, actual_title, actual_meta, h1_count = "ERROR", "", "", "", 0

        except requests.exceptions.SSLError:
            log(f"SSL ERROR : {url}")
            status_code, actual_h1, actual_title, actual_meta, h1_count = "ERROR", "", "", "", 0

        except Exception as e:
            log(f"ERROR : {url} | {e}")
            status_code, actual_h1, actual_title, actual_meta, h1_count = "ERROR", "", "", "", 0

        # This URL is done (success or failure) — clear the transient
        # retry log now that its outcome is captured below and in the
        # final report. Only clear if something was actually shown.
        if log_lines:
            log_box.empty()

        df.at[index, "Status_Code"] = str(status_code)
        df.at[index, "Checked_At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        detail_lines = [
            f"Processing : {index + 1} / {total_urls}",
            f"URL        : {url}",
            f"Status     : {status_code}",
        ]

        h1_result = title_result = meta_result = None

        if check_h1:
            if status_code == "ERROR":
                h1_result = "ERROR"
            elif expected_h1 == "":
                h1_result = "NO EXPECTED H1"
            elif actual_h1 == "":
                h1_result = "H1 MISSING"
            elif h1_count > 1:
                h1_result = "MULTIPLE H1"
            elif expected_h1.lower() == actual_h1.lower():
                h1_result = "PASS"
            else:
                h1_result = "FAIL"

            df.at[index, "Actual_H1"] = actual_h1
            df.at[index, "H1_Count"] = h1_count
            df.at[index, "H1_Result"] = h1_result
            detail_lines.append(f"H1 Result  : {h1_result}  (count: {h1_count})")

        if check_title:
            if status_code == "ERROR":
                title_result = "ERROR"
            elif expected_title == "":
                title_result = "NO EXPECTED TITLE"
            elif actual_title == "":
                title_result = "TITLE MISSING"
            elif expected_title.lower() == actual_title.lower():
                title_result = "PASS"
            else:
                title_result = "FAIL"

            df.at[index, "Actual_Title"] = actual_title
            df.at[index, "Title_Result"] = title_result
            detail_lines.append(f"Title      : {title_result}")

        if check_meta:
            if status_code == "ERROR":
                meta_result = "ERROR"
            elif expected_meta == "":
                meta_result = "NO EXPECTED META"
            elif actual_meta == "":
                meta_result = "META MISSING"
            elif expected_meta.lower() == actual_meta.lower():
                meta_result = "PASS"
            else:
                meta_result = "FAIL"

            df.at[index, "Actual_Meta"] = actual_meta
            df.at[index, "Meta_Result"] = meta_result
            detail_lines.append(f"Meta       : {meta_result}")

        detail_placeholder.code("\n".join(detail_lines))

        row_results = [r for r in [h1_result, title_result, meta_result] if r is not None]
        if "FAIL" in row_results:
            fail_count += 1
        elif status_code == "ERROR":
            error_count += 1
        elif any(r in WARN_VALUES for r in row_results):
            warn_count += 1
        else:
            pass_count += 1

        progress_bar.progress((index + 1) / total_urls)
        time.sleep(delay_between_requests)

    detail_placeholder.code("Done processing all URLs.")

    # ==========================================================
    # BUILD FORMATTED OUTPUT WORKBOOK (in memory)
    # ==========================================================
    excel_buffer = io.BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    wb = load_workbook(excel_buffer)
    ws = wb.active

    col_map = {}
    for cell in ws[1]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    result_columns = [c for c in ["H1_Result", "Title_Result", "Meta_Result"] if c in df.columns]
    status_column = "Status_Code"

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HEADER_BORDER = Border(bottom=Side(style="medium", color="FFFFFF"))

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 3), 16)

    for row_num in range(2, ws.max_row + 1):
        for col_name in result_columns:
            if col_name in col_map:
                cell = ws.cell(row=row_num, column=col_map[col_name])
                apply_cell_style(cell, cell.value)
        if status_column in col_map:
            cell = ws.cell(row=row_num, column=col_map[status_column])
            apply_status_style(cell, cell.value)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    # ==========================================================
    # SUMMARY + DOWNLOAD
    # ==========================================================
    st.success("SEO check completed.")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Passed", pass_count)
    c2.metric("Failed", fail_count)
    c3.metric("Warnings", warn_count)
    c4.metric("Errors", error_count)

    output_filename = f"{source_name}_SEO_Result_001.xlsx"

    st.download_button(
        label="⬇ Download Results (.xlsx)",
        data=output_buffer,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with st.expander("View results table"):
        st.dataframe(df)
